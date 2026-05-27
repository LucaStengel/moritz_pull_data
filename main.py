"""
Preise für Latex (LGM API), Baumwolle Future (Onvista) und Nitril/Butadien (100ppi)
abrufen, in EUR umrechnen und in prices.xlsx speichern.

Verwendung:
  uv run main.py               – Preise abrufen und in prices.xlsx speichern
  uv run main.py -graph        – Graphen für alle drei (Standard: letzte 30 Tage)
  uv run main.py -graph -30d   – Graphen für die letzten 30 Tage
  uv run main.py -graph -7d    – Graphen für die letzten 7 Tage
"""

import re
import sys
from base64 import b64encode
from datetime import datetime, timedelta
from pathlib import Path

import cloudscraper
import matplotlib.dates as mdates
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import openpyxl
import requests
from bs4 import BeautifulSoup
from Visualize.Visualize import Visualize

EXCEL_FILE = Path(__file__).parent / "prices.xlsx"

# Spaltenköpfe der Excel-Datei
COLS = [
    "Datum",
    "Latex (USD/t)",
    "Latex (EUR/t)",
    "Baumwolle (USc/lb)",
    "Baumwolle (EUR/t)",
    "Nitril (CNY/t)",
    "Nitril (EUR/t)",
    "EUR/USD Kurs",
    "EUR/CNY Kurs",
]

# LGM-API: Credentials aus dem öffentlichen Angular-Bundle
_LGM_API = "https://www.lgm.gov.my/webv2api/api/rubberprice/currentprice"
_LGM_AUTH = "Basic " + b64encode(b"FOB:LgMF0b$2025").decode()
_HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}

# 1 USc/lb → USD/t (1 lb = 0.453592 kg → 1 t = 2204.62 lb; 1 USc = 0.01 USD)
_USC_PER_LB_TO_USD_PER_T = 22.0462


# ---------------------------------------------------------------------------
# Wechselkurse
# ---------------------------------------------------------------------------

def fetch_exchange_rates() -> dict[str, float]:
    """Aktuelle EUR-Wechselkurse (kostenlose API, kein Key nötig).

    Gibt zurück: {"USD": <1 EUR in USD>, "CNY": <1 EUR in CNY>}
    """
    try:
        resp = requests.get(
            "https://api.exchangerate-api.com/v4/latest/EUR",
            headers=_HEADERS,
            timeout=10,
        )
        resp.raise_for_status()
        rates = resp.json()["rates"]
        return {"USD": float(rates["USD"]), "CNY": float(rates["CNY"])}
    except Exception as e:
        print(f"  ! Wechselkurs-Abruf fehlgeschlagen ({e}), verwende Näherungswerte")
        return {"USD": 1.08, "CNY": 7.80}


# ---------------------------------------------------------------------------
# Scraper
# ---------------------------------------------------------------------------

def fetch_latex() -> float:
    """Latex in Bulk (USD/t) von der LGM-API."""
    resp = requests.get(
        _LGM_API,
        headers={**_HEADERS, "Authorization": _LGM_AUTH},
        timeout=15,
    )
    resp.raise_for_status()
    for item in resp.json():
        if item.get("grade") == "Latex in Bulk":
            return float(item["sellersUs"])
    raise ValueError("Grade 'Latex in Bulk' nicht in LGM-Antwort gefunden")


def fetch_cotton() -> float:
    """Baumwolle Future aktueller Kurs (USc/lb) von Onvista."""
    url = "https://www.onvista.de/rohstoffe/Baumwolle-Future-21474937"
    resp = requests.get(url, headers=_HEADERS, timeout=15)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")
    tag = soup.find("data", class_=lambda c: c and "text-4xl" in c and "font-bold" in c)
    if not tag or not tag.get("value"):
        raise ValueError("Baumwollpreis-Element nicht auf Onvista gefunden")
    return float(tag["value"])


def fetch_nitrile() -> float:
    """Nitril/Butadien Benchmark-Preis (CNY/t) von 100ppi (Cloudflare via cloudscraper)."""
    url = "https://www.100ppi.com/vane/detail-886.html"
    scraper = cloudscraper.create_scraper()
    resp = scraper.get(url, timeout=20)
    resp.raise_for_status()
    m = re.search(r"基准价[为：:]([\d,]+\.?\d*)元/吨", resp.text)
    if not m:
        m = re.search(r"([\d,]{4,}\.?\d*)元/吨", resp.text)
    if not m:
        raise ValueError("Nitril-Preis-Muster nicht auf 100ppi gefunden")
    return float(m.group(1).replace(",", ""))


# ---------------------------------------------------------------------------
# Umrechnungen
# ---------------------------------------------------------------------------

def to_eur_per_t(value: float, source_unit: str, rates: dict[str, float]) -> float:
    """Rechnet einen Rohstoffpreis in EUR/t um.

    source_unit: "USD/t" | "USc/lb" | "CNY/t"
    """
    if source_unit == "USD/t":
        return round(value / rates["USD"], 2)
    if source_unit == "USc/lb":
        usd_per_t = value * _USC_PER_LB_TO_USD_PER_T
        return round(usd_per_t / rates["USD"], 2)
    if source_unit == "CNY/t":
        return round(value / rates["CNY"], 2)
    raise ValueError(f"Unbekannte Einheit: {source_unit}")


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def _get_workbook():
    if EXCEL_FILE.exists():
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rohstoffpreise"
        ws.append(COLS)
        for col, width in zip("ABCDEFGHI", [20, 14, 12, 17, 15, 13, 12, 11, 11]):
            ws.column_dimensions[col].width = width
        # Kopfzeile fett
        from openpyxl.styles import Font
        for cell in ws[1]:
            cell.font = Font(bold=True)
    return wb, ws


def append_prices(
    latex: float,
    cotton: float,
    nitrile: float,
    rates: dict[str, float],
) -> None:
    latex_eur = to_eur_per_t(latex, "USD/t", rates) if not _isnan(latex) else float("nan")
    cotton_eur = to_eur_per_t(cotton, "USc/lb", rates) if not _isnan(cotton) else float("nan")
    nitrile_eur = to_eur_per_t(nitrile, "CNY/t", rates) if not _isnan(nitrile) else float("nan")

    wb, ws = _get_workbook()
    ws.append([
        datetime.now().replace(microsecond=0),
        latex, latex_eur,
        cotton, cotton_eur,
        nitrile, nitrile_eur,
        round(rates["USD"], 4),
        round(rates["CNY"], 4),
    ])
    wb.save(EXCEL_FILE)

    print()
    print("  Gespeichert in prices.xlsx:")
    print(f"  {'Rohstoff':<12} {'Originalpreis':<20} {'EUR/t'}")
    print("  " + "-" * 50)
    print(f"  {'Latex':<12} {latex} USD/t{'':<12} {latex_eur} EUR/t")
    print(f"  {'Baumwolle':<12} {cotton} USc/lb{'':<11} {cotton_eur} EUR/t")
    print(f"  {'Nitril':<12} {nitrile} CNY/t{'':<12} {nitrile_eur} EUR/t")
    print()
    print(f"  Wechselkurse (Stand heute): 1 EUR = {rates['USD']} USD | 1 EUR = {rates['CNY']} CNY")
    print(f"  Datei: {EXCEL_FILE}")


def _isnan(x: float) -> bool:
    import math
    return math.isnan(x)


# ---------------------------------------------------------------------------
# Graphen
# ---------------------------------------------------------------------------

def show_graphs(days: int) -> None:
    wb, ws = _get_workbook()
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    if not rows:
        print("Keine Daten in der Excel-Datei vorhanden.")
        return

    cutoff = datetime.now() - timedelta(days=days)
    dates = []
    latex_eur, cotton_eur, nitrile_eur = [], [], []
    rate_usd, rate_cny = [], []

    for row in rows:
        if len(row) < 9 or row[0] is None:
            continue
        dt = row[0] if isinstance(row[0], datetime) else datetime.fromisoformat(str(row[0]))
        if dt < cutoff:
            continue
        try:
            dates.append(mdates.date2num(dt))
            latex_eur.append(float(row[2]))
            cotton_eur.append(float(row[4]))
            nitrile_eur.append(float(row[6]))
            rate_usd.append(float(row[7]))
            rate_cny.append(float(row[8]))
        except (TypeError, ValueError):
            continue

    if not dates:
        print(f"Keine Daten in den letzten {days} Tagen vorhanden.")
        return

    x = np.array(dates)

    # 5 feste Positionen: erster Tag, 3 gleichmäßig verteilt, letzter Tag
    x_ticks = np.linspace(x[0], x[-1], 5)

    def _plot(title: str, signals: list[tuple[list, str]], ylabel: str) -> None:
        viz = Visualize(title, y_label=ylabel, x_label="Datum")
        for values, label in signals:
            viz.add_signal(np.array(values, dtype=float), x_axis=x, label=label, scatter=True)

        # Visualize._redraw() ruft cla() auf → löscht Formatter und Locator.
        # Daher _redraw patchen, damit die Datumsformatierung nach jedem Neuzeichnen
        # wiederhergestellt wird.
        _orig = viz._redraw

        def _redraw(n: int) -> None:
            _orig(n)
            viz._ax.xaxis.set_major_locator(mticker.FixedLocator(x_ticks))
            viz._ax.xaxis.set_major_formatter(mdates.DateFormatter("%d.%m.%y"))
            viz._ax.tick_params(axis="x", rotation=45)

        viz._redraw = _redraw
        viz.show()

    # Alle drei Rohstoffe in EUR/t – kombiniertes Diagramm
    _plot(
        f"Rohstoffpreise in EUR/t – letzte {days} Tage",
        [
            (latex_eur,   "Latex (EUR/t)"),
            (cotton_eur,  "Baumwolle (EUR/t)"),
            (nitrile_eur, "Nitril (EUR/t)"),
        ],
        "EUR/t",
    )

    # Wechselkurse als separate Diagramme
    _plot(
        f"Wechselkurs EUR/USD – letzte {days} Tage",
        [(rate_usd, "1 EUR in USD")],
        "USD",
    )
    _plot(
        f"Wechselkurs EUR/CNY – letzte {days} Tage",
        [(rate_cny, "1 EUR in CNY")],
        "CNY",
    )


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _parse_days(argv: list[str]) -> int:
    for arg in argv:
        m = re.match(r"^-(\d+)[dD]$", arg)
        if m:
            return int(m.group(1))
    return 30


def main() -> None:
    argv = sys.argv[1:]

    if "-graph" in argv:
        days = _parse_days(argv)
        print(f"Zeige Graphen für die letzten {days} Tage …")
        show_graphs(days)
        return

    print("Rufe aktuelle Preise und Wechselkurse ab …")

    rates = fetch_exchange_rates()
    print(f"  ✓ Wechselkurse: 1 EUR = {rates['USD']} USD | 1 EUR = {rates['CNY']} CNY")

    latex = cotton = nitrile = float("nan")

    try:
        latex = fetch_latex()
        print(f"  ✓ Latex:     {latex} USD/t  →  {to_eur_per_t(latex, 'USD/t', rates)} EUR/t")
    except Exception as e:
        print(f"  ✗ Latex: {e}")

    try:
        cotton = fetch_cotton()
        print(f"  ✓ Baumwolle: {cotton} USc/lb  →  {to_eur_per_t(cotton, 'USc/lb', rates)} EUR/t")
    except Exception as e:
        print(f"  ✗ Baumwolle: {e}")

    try:
        nitrile = fetch_nitrile()
        print(f"  ✓ Nitril:    {nitrile} CNY/t  →  {to_eur_per_t(nitrile, 'CNY/t', rates)} EUR/t")
    except Exception as e:
        print(f"  ✗ Nitril: {e}")

    if all(_isnan(v) for v in [latex, cotton, nitrile]):
        print("\nAlle Abrufe fehlgeschlagen – nichts gespeichert.")
        sys.exit(1)

    append_prices(latex, cotton, nitrile, rates)


if __name__ == "__main__":
    main()
