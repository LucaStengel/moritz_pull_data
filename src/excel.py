"""Excel-Datei lesen, schreiben und Preise umrechnen."""

import math
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

EXCEL_FILE = Path(__file__).parent.parent / "prices.xlsx"

# Spaltenköpfe – die ersten 9 Spalten bleiben identisch mit dem Ausgangsformat,
# damit bestehende Excel-Zeilen ohne Migration weiterhin korrekt gelesen werden.
COLS = [
    "Datum",
    "Latex (USD/t)",
    "Latex (EUR/t)",
    "Baumwolle (USc/lb)",
    "Baumwolle (EUR/t)",
    "Nitril (CNY/t)",
    "Nitril (EUR/t)",
    "EUR/USD Kurs",
    "EUR/CNY Kurs",
    "Öl WTI (USD/bbl)",
    "Öl WTI (EUR/bbl)",
    "Öl Brent (USD/bbl)",
    "Öl Brent (EUR/bbl)",
    "Erdgas TTF (EUR/MWh)",
    "Polyethylen (CNY/t)",
    "Polyethylen (EUR/t)",
    "USD/MYR Kurs",
    "USD/THB Kurs",
]

_COL_WIDTHS = [
    20, 14, 12, 17, 15, 13, 12, 11, 11,
    15, 13, 16, 14, 17, 15, 13, 12, 12,
]

# 1 USc/lb → USD/t  (1 t = 2204.62 lb, 1 USc = 0.01 USD)
_USC_PER_LB_TO_USD_PER_T = 22.0462


# ---------------------------------------------------------------------------
# Umrechnungen
# ---------------------------------------------------------------------------

def to_eur(value: float, unit: str, rates: dict[str, float]) -> float:
    """Rohstoffpreis in EUR/t (bzw. EUR/bbl) umrechnen.

    Unterstützte Einheiten: 'USD/t', 'USD/bbl', 'USc/lb', 'CNY/t'
    """
    if unit == "USD/t" or unit == "USD/bbl":
        return round(value / rates["USD"], 2)
    if unit == "USc/lb":
        return round(value * _USC_PER_LB_TO_USD_PER_T / rates["USD"], 2)
    if unit == "CNY/t":
        return round(value / rates["CNY"], 2)
    raise ValueError(f"Unbekannte Einheit: {unit}")


def isnan(x: float) -> bool:
    return math.isnan(x)


def safe_convert(value: float, unit: str, rates: dict[str, float]) -> float:
    """Wie to_eur(), gibt aber NaN zurück statt zu rechnen wenn value NaN ist."""
    return to_eur(value, unit, rates) if not isnan(value) else float("nan")


# ---------------------------------------------------------------------------
# Workbook-Verwaltung
# ---------------------------------------------------------------------------

def get_workbook():
    """Öffnet oder erstellt die Excel-Datei und ergänzt fehlende Spaltenköpfe."""
    if EXCEL_FILE.exists():
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        for col_idx, name in enumerate(COLS, 1):
            if ws.cell(1, col_idx).value is None:
                cell = ws.cell(1, col_idx, name)
                cell.font = Font(bold=True)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rohstoffpreise"
        ws.append(COLS)
        for cell in ws[1]:
            cell.font = Font(bold=True)

    for col_idx, width in enumerate(_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    return wb, ws


# ---------------------------------------------------------------------------
# Daten speichern
# ---------------------------------------------------------------------------

def append_prices(
    latex: float,
    cotton: float,
    nitrile: float,
    oil_wti: float,
    oil_brent: float,
    gas_ttf: float,
    polyethylene: float,
    rates: dict[str, float],
) -> None:
    """Aktuelle Preise als neue Zeile in die Excel-Datei schreiben und Ergebnis ausgeben."""
    latex_eur     = safe_convert(latex,       "USD/t",   rates)
    cotton_eur    = safe_convert(cotton,      "USc/lb",  rates)
    nitrile_eur   = safe_convert(nitrile,     "CNY/t",   rates)
    oil_wti_eur   = safe_convert(oil_wti,     "USD/bbl", rates)
    oil_brent_eur = safe_convert(oil_brent,   "USD/bbl", rates)
    pe_eur        = safe_convert(polyethylene,"CNY/t",   rates)

    # Kreuzrate: USD/MYR = (1 EUR in MYR) / (1 EUR in USD)
    usd_myr = round(rates["MYR"] / rates["USD"], 4)
    usd_thb = round(rates["THB"] / rates["USD"], 4)

    wb, ws = get_workbook()
    ws.append([
        datetime.now().replace(microsecond=0),
        latex,       latex_eur,
        cotton,      cotton_eur,
        nitrile,     nitrile_eur,
        round(rates["USD"], 4),
        round(rates["CNY"], 4),
        oil_wti,     oil_wti_eur,
        oil_brent,   oil_brent_eur,
        gas_ttf,
        polyethylene, pe_eur,
        usd_myr,
        usd_thb,
    ])
    wb.save(EXCEL_FILE)

    def fmt(v, unit=""):
        return f"{v} {unit}".strip() if not isnan(v) else "–"

    print()
    print("  Gespeichert in prices.xlsx:")
    print(f"  {'Rohstoff':<16} {'Originalpreis':<22} {'EUR-Preis'}")
    print("  " + "-" * 60)
    print(f"  {'Latex':<16} {fmt(latex, 'USD/t'):<22} {fmt(latex_eur, 'EUR/t')}")
    print(f"  {'Baumwolle':<16} {fmt(cotton, 'USc/lb'):<22} {fmt(cotton_eur, 'EUR/t')}")
    print(f"  {'Nitril':<16} {fmt(nitrile, 'CNY/t'):<22} {fmt(nitrile_eur, 'EUR/t')}")
    print(f"  {'Öl WTI':<16} {fmt(oil_wti, 'USD/bbl'):<22} {fmt(oil_wti_eur, 'EUR/bbl')}")
    print(f"  {'Öl Brent':<16} {fmt(oil_brent, 'USD/bbl'):<22} {fmt(oil_brent_eur, 'EUR/bbl')}")
    print(f"  {'Erdgas TTF':<16} {fmt(gas_ttf, 'EUR/MWh'):<22}")
    print(f"  {'Polyethylen':<16} {fmt(polyethylene, 'CNY/t'):<22} {fmt(pe_eur, 'EUR/t')}")
    print()
    print(f"  Wechselkurse: 1 EUR = {rates['USD']:.4f} USD  |  1 EUR = {rates['CNY']:.4f} CNY")
    print(f"                1 USD = {usd_myr:.4f} MYR   |  1 USD = {usd_thb:.4f} THB")
    print(f"  Datei: {EXCEL_FILE}")
